# -*- coding: utf-8 -*-
"""
@author: jmartorell
"""
import os
import ssl
import time
import xlrd
import json
import smtplib
import operator
import targetX as t
import dataset as d
import RESTful_api as api
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.firefox import GeckoDriverManager
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Side, Border
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from openpyxl import load_workbook
from openpyxl import Workbook
from bs4 import BeautifulSoup
from email import encoders


def send_attachment():
    with open('../../../Documents/keys.json', 'r') as a:
        keys_dict = json.loads(a.read())
    subject = "An email with attachment from Python"
    body = "This is an email with attachment sent from Python"
    sender_email = keys_dict['mailAddress'][0]
    receiver_email = keys_dict['mailAddress'][1]
    # password = input("Type your password and press enter:")
    password = keys_dict['mailPassword'][0]
    a.close()

    # Create a multipart message and set headers
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message["Bcc"] = receiver_email  # Recommended for mass emails

    # Add body to email
    message.attach(MIMEText(body, "plain"))

    spreadsheet = '//home/jmartorell/Booking/bookings/bookings.xlsx'
    filename = spreadsheet  # In same directory as script

    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()

    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, text)
        print('Sending email ...')


class App:
    # 1) init method initializes variables that will be accessible by self from any method of the class
    def __init__(self, keys='../../../Documents/keys.json', target_destination='New york', depart_m='2', depart_w='3',
                 depart_d='1', return_m='2', return_w='3', return_d='7', cell_city='New York', cell_cc='US',
                 path='../../../Booking'):
        self.keys = keys
        self.target_destination = target_destination
        self.depart_m = depart_m
        self.depart_w = depart_w
        self.depart_d = depart_d
        self.return_m = return_m
        self.return_w = return_w
        self.return_d = return_d
        self.cell_city = cell_city
        self.cell_cc = cell_cc
        self.path = path
        self.browser = webdriver.Firefox(executable_path=GeckoDriverManager().install())
        self.error = False
        self.url = 'https://pro.w2m.travel'
        self.all_hotels = []
        self.all_prices = []
        self.all_locations = []
        self.all_positions = []
        self.display = []
        self.cheap = []
        self.data = {}
        self.index = ""
        self.euro_symbol = '€'
        self.coefficient = '1.374'
        self.target_recharge = ""
        self.browser.get(self.url)
        self.log_in()
        if self.error is False:
            self.search_engine_insert()
        if self.error is False:
            self.scroll_down()
        if self.error is False:
            if not os.path.exists(path):
                os.mkdir(path)
            self.file_manager()
            # todo: self.reach_target()
        # close the browser
        sleep(1)
        self.browser.quit()

    # 2) log in method allows us to log in to access the provider's services
    def log_in(self, ):
        try:
            with open(self.keys, 'r') as a:
                keys_dict = json.loads(a.read())
            input_element = self.browser.find_element_by_id("email")
            input_element.clear()
            input_element.send_keys(keys_dict['username'][3])
            input_element = self.browser.find_element_by_id("password")
            input_element.clear()
            print('Logging in with username and password ...')
            input_element.send_keys(keys_dict['password'][3])
            input_element.submit()
            a.close()

            # print(self.browser.current_url)
        except Exception:
            print('Some exception occurred while trying to find username or password field')
            self.error = True

    def cookies_popup(self):
        print('Closing cookies window ...')
        WebDriverWait(self.browser, 100).until(EC.visibility_of_element_located((
            By.CSS_SELECTOR,
            '.cookie-policy__close'))).click()
        sleep(1)

    # 3) search engine insert method fills in the search engine fields and clicks on the search button
    def search_engine_insert(self):
        # wait to load the search engine
        WebDriverWait(self.browser, 100).until(EC.visibility_of_element_located((
            By.XPATH,
            '//*[@id="hotel-searcher-_ctl1__ctl1__ctl1_pageBody_pageBody_searcher__ctl0_ctlZoneSelector-input"]'))).click()
        self.cookies_popup()
        element = self.browser.find_element_by_xpath(
            '//*[@id="hotel-searcher-_ctl1__ctl1__ctl1_pageBody_pageBody_searcher__ctl0_ctlZoneSelector-input"]')
        element.clear()

        # check access
        # assert "Hoteles | W2M" in self.browser.title

        # enter data in input field
        element.send_keys(self.target_destination)

        # TODO:
        # drop-down item selection
        actions = ActionChains(self.browser)
        for _ in range(1):
            actions.send_keys(Keys.ARROW_DOWN).perform()
            sleep(1)

        # enter destination city
        target_city = element.find_element_by_xpath(
            "//div[3]/div[1]")
        target_city.click()

        # press the search button
        login_attempt = element.find_element_by_xpath(
            "//div[2]/div[2]/button")
        print('Loading page ...')
        login_attempt.click()

    # 4) the reach target method systematically selects the first objective by clicking on it
    # todo: def reach_target(self):
    def reach_target(self, index):
        target_button = self.browser.find_element_by_xpath(
            '//div[ ' + index + ' ]/article/div[1]/div[2]/div[2]/div/div[2]/span/a')
        self.browser.execute_script("arguments[0].scrollIntoView();", target_button)
        # target_button.click()

    def scroll_down(self):
        self.browser.implicitly_wait(15)

        # todo REF: https://stackoverflow.com/questions/48006078/how-to-scroll-down-in-python-selenium-step-by-step
        read_mores = self.browser.find_elements_by_xpath('//div[text()="Best stay price"]')
        print('Scrolling page ...')
        for read_more in read_mores:
            self.browser.execute_script("arguments[0].scrollIntoView();", read_more)
            # read_more.click()

        print("Scraping page ...")
        soup = BeautifulSoup(self.browser.page_source, 'lxml')
        hotel_list = soup.find_all('div', {'class': 'results-list__item'})

        print("\n\tDisplay:\n")
        try:
            for i, hotel in enumerate(hotel_list):
                self.all_positions.append(i + 1)
                hotel_name = hotel.find('h2', {'class': 'info-card__title'}).getText()
                # fixme: remove whitespaces REF: https://stackoverrun.com/es/q/743639
                hotel_name = ' '.join(hotel_name.split())
                # notice that instead of .getText().strip('€') here we work with .getText().replace('€', ''):
                hotel_price = hotel.find('a', {'tabindex': '0'}).getText().replace('€', '')
                hotel_price = hotel_price.replace('.', '')
                hotel_price = hotel_price.replace(',', '.')
                hotel_price = float(hotel_price)
                hotel_price = "{0:.2f}".format(hotel_price)
                self.all_prices.append(hotel_price)

                hotel_location = hotel.find('div', {'class': 'info-card__location'}).getText().strip(',')
                hotel_location = ' '.join(hotel_location.split()).rstrip("View the map")
                self.all_locations.append(hotel_location)

                if len(hotel_price) == 5:
                    hotel_price = "   " + hotel_price
                if len(hotel_price) == 6:
                    hotel_price = "  " + hotel_price
                if len(hotel_price) == 7:
                    hotel_price = " " + hotel_price
                if len(hotel_price) == 8:
                    hotel_price = "" + hotel_price
                self.all_hotels.append(hotel_name)
                if i < 9:
                    print(" %d - %s %s %s - %s" % (i + 1, hotel_price, self.euro_symbol, hotel_name, hotel_location))
                else:
                    print("%d - %s %s %s - %s" % (i + 1, hotel_price, self.euro_symbol, hotel_name, hotel_location))

            print("\n\tRanking:\n")
            # float cast
            new_prices = []
            for element in self.all_prices:
                rank = float(element)
                new_prices.append(rank)

            # final list
            display_list = zip(self.all_positions, self.all_hotels, new_prices, self.all_locations)
            ranking = sorted(display_list, key=operator.itemgetter(2))
            for j, k, v, w in ranking:
                if v < 100.00:
                    print("   ", "{0:.2f}".format(v), k)
                if 99.00 < v < 1000.00:
                    print("  ", "{0:.2f}".format(v), k)
                if 999.00 < v < 10000.00:
                    print(" ", "{0:.2f}".format(v), k)
                if v > 9999.00:
                    print("", "{0:.2f}".format(v), k)

            self.display = display_list
            self.data = ranking
            self.cheap = ranking[0]
            print('\nLocated booking for', self.cheap[2], self.euro_symbol, '...')
            print('Pointing to the target button', self.cheap[0], '...')
            self.index = str(self.cheap[0])
            if self.error is False:
                self.reach_target(self.index)

            sleep(2)
        except Exception as e:
            self.error = True
            print(e)
            print('Some error occurred while trying to scroll down')

    def set_stylesheet(self, sheet):

        # time frame:
        sheet.merge_cells('A1:L1')
        time_frame = sheet['A1']
        time_frame.fill = PatternFill(
            start_color="0007147A", end_color="0007147A", fill_type="solid")
        time_frame.font = Font(bold=True, size=11)
        bd = Side(style='thick', color="000000")
        time_frame.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        # timestamp
        time_label = 'Business&Travel:        %s                       Time Frame:        %s%s/2020  -  %s%s/2020' \
                     % (time.ctime(), t.dep + "/", t.start_month, t.ret + "/", t.end_month)
        sheet.cell(row=1, column=1).value = time_label

        # title bar
        # fixme REF:
        # https://stackoverflow.com/questions/35918504/adding-a-background-color-to-cell-openpyxl
        for col_range in range(1, 13):
            cell_title = sheet.cell(2, col_range)
            cell_title.fill = PatternFill(start_color="00f4f4f7", end_color="00f4f4f7", fill_type="solid")
            cell_title = sheet.cell(2, col_range)
            cell_title.font = Font(bold=True, size=11)
            bd = Side(style='thick', color="000000")
            cell_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        header = ('Code', 'Price', 'Retail', 'Profit', 'CC', 'City', 'No', 'Hotel', 'Hu', 'Gr', 'Co', 'Location')
        sheet.cell(row=2, column=1).value = header[0]
        sheet.cell(row=2, column=2).value = header[1]
        sheet.cell(row=2, column=3).value = header[2]
        sheet.cell(row=2, column=4).value = header[3]
        sheet.cell(row=2, column=5).value = header[4]
        sheet.cell(row=2, column=6).value = header[5]
        sheet.cell(row=2, column=7).value = header[6]
        sheet.cell(row=2, column=8).value = header[7]
        sheet.cell(row=2, column=9).value = header[8]
        sheet.cell(row=2, column=10).value = header[9]
        sheet.cell(row=2, column=11).value = header[10]
        sheet.cell(row=2, column=12).value = header[11]

        sheet.column_dimensions['B'].number_format = '#,##0.00'
        sheet.column_dimensions['C'].number_format = '#,##0.00'
        sheet.column_dimensions['D'].number_format = '#,##0.00'
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 9
        sheet.column_dimensions['C'].width = 9
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 4
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 4
        sheet.column_dimensions['H'].width = 60
        sheet.column_dimensions['I'].width = 4
        sheet.column_dimensions['J'].width = 4
        sheet.column_dimensions['K'].width = 4
        sheet.column_dimensions['L'].width = 50

        format = sheet.column_dimensions['A']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['B']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['C']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['D']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['E']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['F']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['G']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['H']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['I']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['J']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['K']
        format.font = Font(bold=True, italic=True, name='Arial')
        format = sheet.column_dimensions['L']
        format.font = Font(bold=True, italic=True, name='Arial')

    def read_code(self):
        global trip_code
        f = open("trip_code.txt", "r")
        if f.mode == 'r':
            trip_code = f.read()
        return trip_code

    def write_code(self, input_code):
        f = open("trip_code.txt", "w")
        f.write(input_code)
        f.close()

    def read_bookings_from_excel_file(self, excel_path):
        workbook = xlrd.open_workbook(excel_path)
        worksheet = workbook.sheet_by_index(0)
        cell_h3 = worksheet.cell_value(2, 7)
        return cell_h3

    def write_bookings_to_excel_file(self, booking_path):

        # FIXME: openpyxl -> https://openpyxl.readthedocs.io/en/stable/index.html
        filepath = os.path.join(booking_path, 'bookings.xlsx')
        print('Writing to excel ...')
        if not os.path.exists(filepath):
            workbook = Workbook()
            workbook.save(filepath)
            workbook.create_sheet("Spapshoot", 0)
            workbook.create_sheet("Display", 1)
        else:
            workbook = load_workbook(filepath)

        # fixme: delete the default sheet:
        if "Sheet" in workbook.sheetnames:
            std = workbook["Sheet"]
            workbook.remove(std)

        # switch sheet
        workbook.active = 0

        sheet = workbook.active
        self.set_stylesheet(sheet)

        # write sheet
        i = 3
        w = 0
        for row in self.data:
            cell_reference = sheet.cell(row=i, column=1)
            update_code = t.code_builder(self.read_code())
            self.write_code(update_code)
            cell_reference.value = update_code
            cell_reference = sheet.cell(row=i, column=2)
            cell_reference.value = row[2]
            sheet['C{}'.format(i)] = '=PRODUCT(B{},{}'.format(i, self.coefficient)
            sheet['D{}'.format(i)] = '=SUM(C{},-B{}'.format(i, i)
            cell_reference = sheet.cell(row=i, column=5)
            cell_reference.value = self.cell_cc
            cell_reference = sheet.cell(row=i, column=6)
            cell_reference.value = self.cell_city
            cell_reference = sheet.cell(row=i, column=7)
            cell_reference.value = row[0]
            cell_reference = sheet.cell(row=i, column=8)
            cell_reference.value = row[1]
            # weather rates
            humidity = api.get_humidity(d.tour_en[w][0])
            cell_reference = sheet.cell(row=i, column=9)
            cell_reference.value = humidity
            grades = api.get_temperature(d.tour_en[w][0])
            cell_reference = sheet.cell(row=i, column=10)
            cell_reference.value = grades

            cell_reference = sheet.cell(row=i, column=1)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=2)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=3)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=4)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=5)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=6)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=7)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=8)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=9)
            bd = Side(style='thick', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=10)
            bd = Side(style='thick', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=11)
            bd = Side(style='thick', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell_reference = sheet.cell(row=i, column=12)
            bd = Side(style='thin', color="000000")
            cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)

            if grades > 19:
                covid_green = PatternFill(
                    start_color='0000FF00',
                    end_color='0000FF00',
                    fill_type='solid')
                sheet.cell(row=i, column=11).fill = covid_green
            if 20 > grades > 14:
                covid_blue = PatternFill(
                    start_color='000000FF',
                    end_color='000000FF',
                    fill_type='solid')
                sheet.cell(row=i, column=11).fill = covid_blue
            else:
                covid = PatternFill(
                    start_color='00FF0000',
                    end_color='00FF0000',
                    fill_type='solid')
                sheet.cell(row=i, column=11).fill = covid
            cell_reference = sheet.cell(row=i, column=12)
            cell_reference.value = row[3]
            i += 1
        # todo START recharger
        workbook.active = 1
        display_sheet = workbook.active
        self.set_stylesheet(display_sheet)

        # automatic recharger
        target = 3
        while display_sheet.cell(row=target, column=8).value is not None:
            target += 1

        recharger = self.data[0]
        cell_reference = display_sheet.cell(row=target, column=1)
        update_code = t.code_builder(self.read_code())
        self.write_code(update_code)
        cell_reference.value = update_code
        cell_reference = display_sheet.cell(row=target, column=2)
        cell_reference.value = recharger[2]
        display_sheet['C{}'.format(target)] = '=PRODUCT(B{},{}'.format(target, self.coefficient)
        display_sheet['D{}'.format(target)] = '=SUM(C{},-B{}'.format(target, target)
        cell_reference = display_sheet.cell(row=target, column=5)
        cell_reference.value = self.cell_cc
        cell_reference = display_sheet.cell(row=target, column=6)
        cell_reference.value = self.cell_city
        cell_reference = display_sheet.cell(row=target, column=7)
        cell_reference.value = recharger[0]
        cell_reference = display_sheet.cell(row=target, column=8)
        cell_reference.value = recharger[1]
        # weather rates
        humidity = api.get_humidity(d.tour_en[w][0])
        cell_reference = display_sheet.cell(row=target, column=9)
        cell_reference.value = humidity
        grades = api.get_temperature(d.tour_en[w][0])
        cell_reference = display_sheet.cell(row=target, column=10)
        cell_reference.value = grades

        cell_reference = display_sheet.cell(row=target, column=1)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=2)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=3)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=4)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=5)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=6)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=7)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=8)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=9)
        bd = Side(style='thick', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=10)
        bd = Side(style='thick', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=11)
        bd = Side(style='thick', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell_reference = display_sheet.cell(row=target, column=12)
        bd = Side(style='thin', color="000000")
        cell_reference.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        if grades > 19:
            covid_green = PatternFill(
                start_color='0000FF00',
                end_color='0000FF00',
                fill_type='solid')
            display_sheet.cell(row=target, column=11).fill = covid_green
        if 20 > grades > 14:
            covid_blue = PatternFill(
                start_color='000000FF',
                end_color='000000FF',
                fill_type='solid')
            display_sheet.cell(row=target, column=11).fill = covid_blue
        else:
            covid_red = PatternFill(
                start_color='00FF0000',
                end_color='00FF0000',
                fill_type='solid')
            display_sheet.cell(row=target, column=11).fill = covid_red
        cell_reference = display_sheet.cell(row=target, column=12)
        cell_reference.value = recharger[3]

        # switch sheet
        workbook.active = 0
        # todo STOP recharger

        workbook.save(filepath)  # save file

    def file_manager(self, ):
        f = open("trip_code.txt", "w+")
        f.write("LM30")
        f.close()
        bookings_folder_path = os.path.join(self.path, 'bookings')
        if not os.path.exists(bookings_folder_path):
            os.mkdir(bookings_folder_path)
        if self.error is False:
            self.write_bookings_to_excel_file(bookings_folder_path)
        if self.error is False:
            self.target_recharge = self.read_bookings_from_excel_file(self.path + '/bookings/bookings.xlsx')
        print("Writing reservation", self.target_recharge, "in sheet Display ...")


if __name__ == '__main__':
    switch = t.switch
    if switch != 0:
        x = 0
        while x < 24:
            app = App(depart_m=t.depart_month,
                      depart_w=t.depart_week,
                      depart_d=t.depart_day,
                      return_m=t.return_month,
                      return_w=t.return_week,
                      return_d=t.return_day,
                      target_destination=d.tour_en[x][0],
                      cell_city=d.tour_en[x][0],
                      cell_cc=d.tour_en[x][1]
                      )
            x += 1
        send_attachment()
    else:
        app = App(depart_m=t.depart_month,
                  depart_w=t.depart_week,
                  depart_d=t.depart_day,
                  return_m=t.return_month,
                  return_w=t.return_week,
                  return_d=t.return_day,
                  )
